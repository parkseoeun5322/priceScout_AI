"""레이어 2 — 실행 워커.

네이버 검색(쇼핑) 오픈 API를 호출해 상품 후보를 가져오고(단계 1),
검색어 토큰 매칭으로 동일상품을 선별(단계 2)한 뒤, 오케스트레이터(레이어 1)가
`Worker` 타입으로 주입받아 수집한다.

  - 단계 1: API 호출    → clean_title, search_naver
  - 단계 2: 토큰 매칭   → tokenize, match_score, select_best, make_worker
  - 결과 저장(openpyxl) → (예정)

⚠️ Playwright 접근은 네이버 봇 차단으로 폐기됨. (plan/init.md 참조)
"""

from __future__ import annotations

import asyncio
import html
import re
from pathlib import Path
from typing import Awaitable, Callable

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config

# 네이버 API title은 검색어 일치 부분을 <b>...</b>로 감싸고 HTML 엔티티(&amp; 등)를 쓴다.
_TAG_RE = re.compile(r"<[^>]+>")

# 오케스트레이터와 동일한 Worker 타입 (순환 import 방지를 위해 여기서도 선언)
Worker = Callable[[dict], Awaitable[dict]]


# ===========================================================================
# 단계 1 — API 호출
#   네이버 검색 오픈 API(GET /v1/search/shop.json)를 호출해
#   가격 오름차순(sort=asc) 후보 리스트를 정규화해 반환한다.
# ===========================================================================
def clean_title(raw: str) -> str:
    """API title의 <b> 태그를 제거하고 HTML 엔티티를 원문으로 되돌린다.

    예: '<b>스카치</b> 테이프 &amp; 커터' → '스카치 테이프 & 커터'
    """
    if not raw:
        return ""
    return html.unescape(_TAG_RE.sub("", raw)).strip()


def _to_price(value) -> int | None:
    """lprice 등 가격 문자열을 int로. 빈 값/파싱 실패 시 None."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _normalize_item(item: dict) -> dict:
    """오픈 API items[]의 한 항목을 우리가 쓰는 필드로 정규화한다."""
    return {
        "title": clean_title(item.get("title", "")),
        "link": item.get("link", ""),
        "lprice": _to_price(item.get("lprice")),
        "mallName": item.get("mallName", ""),
        "productId": item.get("productId", ""),
        "brand": item.get("brand", ""),
        "maker": item.get("maker", ""),
        "category1": item.get("category1", ""),
        "category2": item.get("category2", ""),
        "category3": item.get("category3", ""),
        "category4": item.get("category4", ""),
    }


async def search_naver(
    query: str,
    *,
    display: int = config.NAVER_DISPLAY,
    sort: str = config.NAVER_SORT,
    client: httpx.AsyncClient | None = None,
    timeout: float = 10.0,
) -> list[dict]:
    """네이버 검색 오픈 API를 호출해 정규화된 상품 후보 리스트를 반환한다.

    - 엔드포인트: GET config.NAVER_SHOP_API_URL
    - 헤더: config.naver_api_headers() (X-Naver-Client-Id / X-Naver-Client-Secret)
    - 파라미터: query, display, sort=asc (가격 오름차순 = 낮은 가격순)

    `client`를 넘기면 그 AsyncClient를 재사용한다(여러 요청에서 커넥션 풀 공유).
    넘기지 않으면 호출마다 임시 클라이언트를 만들어 닫는다.

    HTTP 오류는 httpx 예외로 그대로 전파한다(오케스트레이터의 _process_one이 격리).
    반환: _normalize_item()으로 정규화된 dict 리스트. sort=asc이므로 가격 오름차순.
    """
    params = {"query": query, "display": display, "sort": sort}
    headers = config.naver_api_headers()

    async def _request(c: httpx.AsyncClient) -> list[dict]:
        resp = await c.get(
            config.NAVER_SHOP_API_URL,
            params=params,
            headers=headers,
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        return [_normalize_item(it) for it in data.get("items", [])]

    if client is not None:
        return await _request(client)
    async with httpx.AsyncClient() as c:
        return await _request(c)


# ===========================================================================
# 단계 2 — 토큰 매칭 (동일상품 선별)
#   검색어 토큰이 title에 많이 포함된 후보 중 lprice 최소 항목 채택.
#   items는 sort=asc로 이미 가격 오름차순이므로, 후보 중 첫 번째가 최저가.
# ===========================================================================
def tokenize(text: str) -> list[str]:
    """텍스트를 소문자 토큰 리스트로 분리한다.

    공백 기준으로 나누고 소문자 변환. 빈 토큰은 제외.
    """
    return [t.lower() for t in text.split() if t]


def match_score(query_tokens: list[str], title: str) -> float:
    """쿼리 토큰 중 title에 포함된 비율(0.0~1.0)을 반환한다.

    대소문자 무시 substring 검색: '18mm' 토큰은 title의 '18MM'과 매칭됨.
    query_tokens가 비면 0.0 반환.
    """
    if not query_tokens:
        return 0.0
    title_lower = title.lower()
    matched = sum(1 for t in query_tokens if t in title_lower)
    return matched / len(query_tokens)


def select_best(
    query: str,
    items: list[dict],
    *,
    threshold: float = config.MATCH_THRESHOLD,
) -> tuple[dict | None, str]:
    """토큰 매칭으로 가장 적합한 상품을 선택한다.

    items는 lprice 오름차순(sort=asc)으로 이미 정렬된 상태.
    threshold 이상 매칭된 항목 중 첫 번째(= 최저가)를 채택한다.

    반환: (선택된 item dict 또는 None, 사유 문자열)
      - 매칭 성공: (item, "매칭 67% (5후보 중 최저가)")
      - 후보 없음:  (None, "임계값 미달 (최고 매칭 33%, threshold=50%)")
      - 결과 없음:  (None, "API 결과 없음")
    """
    if not items:
        return None, "API 결과 없음"

    query_tokens = tokenize(query)
    if not query_tokens:
        # 검색어 자체가 비어있으면 최저가 그대로 반환
        return items[0], "검색어 토큰 없음 → 최저가 채택"

    # threshold 이상인 후보만 추림 (이미 가격 오름차순이므로 첫 번째가 최저가)
    candidates = [
        (it, match_score(query_tokens, it["title"]))
        for it in items
        if match_score(query_tokens, it["title"]) >= threshold
    ]

    if candidates:
        best, score = candidates[0]
        return best, f"매칭 {score:.0%} ({len(candidates)}후보 중 최저가)"

    # 임계값 미달이지만 API 결과가 있으면 최저가(첫 번째)를 저신뢰로 채택.
    # → 상태 칸에 "저신뢰 채택:" 표시되므로 사람이 검토할 수 있음.
    best_score = max(match_score(query_tokens, it["title"]) for it in items)
    return items[0], (
        f"저신뢰 채택: 최고 매칭 {best_score:.0%} (threshold={threshold:.0%})"
    )


_PARENS_RE = re.compile(r"\s*\([^)]*\)")


def strip_parens(text: str) -> str:
    """상품명에서 소괄호 및 그 안의 내용을 제거한다.

    예: "클레이 도구, 찍기틀 세트(2종 택1/롤러/모양틀/KC인증)" → "클레이 도구, 찍기틀 세트"
    괄호가 없으면 원문 그대로 반환.
    """
    return _PARENS_RE.sub("", text).strip()


def make_worker(client: httpx.AsyncClient | None = None) -> Worker:
    """네이버 오픈 API 호출(단계 1) + 토큰 매칭(단계 2)을 조합한 Worker를 반환한다.

    `client`를 넘기면 여러 호출에서 커넥션 풀을 공유한다(권장).
    None이면 호출마다 임시 클라이언트를 생성한다.

    반환된 worker(product) → dict 구조:
      - 매칭 성공:   status='ok',       lprice(int), title, link, mallName, match_reason
      - 저신뢰 채택: status='low_match', lprice(int), title, link, mallName, match_reason("저신뢰 채택: ...")
      - API 결과 없음: status='no_match', lprice=None, title=None, link=None, match_reason
      - 예외 발생:   오케스트레이터 _process_one이 status='error'로 격리.

    폴백 순서 (API 0건일 때만 다음 단계로):
      1) "상품명 규격" → 2) "상품명" (규격 있을 때) → 3) 괄호 제거 상품명
    재검색으로 채택된 결과는 match_reason 앞에 "[상품명 재검색]" 또는 "[괄호제거 재검색]" 이 붙는다.
    """
    async def worker(product: dict) -> dict:
        query = product["search_query"]
        items = await search_naver(query, client=client)
        fallback_label: str | None = None

        # 폴백 1: "상품명 규격" 0건 + 규격 있음 → "상품명"만으로 재검색
        if not items and product.get("spec"):
            name_query = product["name"]
            items = await search_naver(name_query, client=client)
            if items:
                fallback_label = "상품명 재검색"
                query = name_query

        # 폴백 2: 여전히 0건 + 상품명에 괄호 있음 → 괄호 제거 후 재검색
        if not items:
            name_no_parens = strip_parens(product["name"])
            if name_no_parens and name_no_parens != product["name"]:
                items = await search_naver(name_no_parens, client=client)
                if items:
                    fallback_label = "괄호제거 재검색"
                    query = name_no_parens

        best, reason = select_best(query, items)

        if best is None:
            return {
                "status": "no_match",
                "match_reason": reason,
                "lprice": None,
                "title": None,
                "link": None,
                "mallName": None,
            }

        is_low = reason.startswith("저신뢰 채택:")
        if fallback_label:
            reason = f"[{fallback_label}] {reason}"
        return {
            "status": "low_match" if is_low else "ok",
            "match_reason": reason,
            "lprice": best["lprice"],
            "title": best["title"],
            "link": best["link"],
            "mallName": best["mallName"],
        }

    return worker


# ===========================================================================
# 결과 저장 (openpyxl)
#   distribute()가 반환한 결과 리스트를 새 워크북에 기록한다.
#   배송비는 "확인필요" 고정, 합계는 qty×lprice(배송비 미포함).
# ===========================================================================

# 출력 시트 헤더 (순서가 열 순서)
_HEADERS = [
    "순번", "상품명", "규격", "수량", "검색어",
    "최저가(단가)", "배송비", "합계",
    "매칭상품명", "상품URL", "쇼핑몰", "상태",
]

# 헤더 배경색 (연한 파란색)
_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def _status_label(result: dict) -> str:
    """결과 dict에서 상태 열에 표시할 문자열을 만든다."""
    status = result.get("status", "")
    reason = result.get("match_reason", "")
    error  = result.get("error", "")
    if status in ("ok", "low_match"):
        return reason   # reason 자체에 "저신뢰 채택:" 등 맥락이 포함되어 있음
    if status == "no_match":
        return f"미매칭: {reason}"
    if status == "error":
        return f"오류: {error}"
    return status


def save_results(results: list[dict], path: Path) -> None:
    """검색 결과 리스트를 엑셀 파일로 저장한다.

    - `path` 의 부모 디렉토리가 없으면 자동 생성.
    - 기존 파일이 있으면 덮어쓴다.
    - 배송비 칸은 "확인필요" 고정.
    - 합계 = 수량 × 최저가(단가). qty 또는 lprice 중 하나라도 None이면 빈칸.
    - 상품URL은 하이퍼링크로 설정.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "최저가목록"

    # 헤더 행
    ws.append(_HEADERS)
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 행
    url_col_idx = _HEADERS.index("상품URL") + 1  # 1-base

    for r in results:
        lprice = r.get("lprice")
        qty    = r.get("qty")
        total  = (qty * lprice) if (isinstance(qty, int) and isinstance(lprice, int)) else None

        row_values = [
            r.get("no", ""),
            r.get("name", ""),
            r.get("spec", ""),
            qty if qty is not None else "",
            r.get("search_query", ""),
            lprice if lprice is not None else "",
            "확인필요",
            total if total is not None else "",
            r.get("title") or "",
            r.get("link") or "",
            r.get("mallName") or "",
            _status_label(r),
        ]
        ws.append(row_values)

        # 상품URL 셀에 하이퍼링크 적용
        link = r.get("link") or ""
        if link:
            cell = ws.cell(row=ws.max_row, column=url_col_idx)
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")

    # 열 너비 자동 조정 (헤더+데이터 최대 길이 기준, 최대 60자)
    for col_idx, header in enumerate(_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(path)
    print(f"[save_results] {len(results)}건 저장 완료 → {path}")


# ===========================================================================
# 진입점 — 단계 1·2 단독 검증용 CLI
#   임의 검색어로 오픈 API를 호출해 토큰 매칭까지 수행한 결과를 출력한다.
#   예: python search_worker.py "스카치 테이프 18mm" --display 10
# ===========================================================================
def main() -> None:
    import argparse

    config.setup_utf8_output()

    parser = argparse.ArgumentParser(
        description="레이어 2: 네이버 검색 오픈 API 호출 + 토큰 매칭 결과를 출력한다."
    )
    parser.add_argument("query", help="검색어 (예: '스카치 테이프 18mm')")
    parser.add_argument(
        "--display", type=int, default=config.NAVER_DISPLAY, help="가져올 결과 수"
    )
    parser.add_argument(
        "--threshold", type=float, default=config.MATCH_THRESHOLD,
        help="토큰 매칭 임계값 (기본 0.5)",
    )
    args = parser.parse_args()

    async def _run():
        print(
            f"검색어: {args.query!r}  display={args.display}  "
            f"sort={config.NAVER_SORT}  threshold={args.threshold:.0%}"
        )
        items = await search_naver(args.query, display=args.display)
        print(f"\n[단계 1] API 결과 {len(items)}건 (가격 오름차순)")
        query_tokens = tokenize(args.query)
        for i, it in enumerate(items, start=1):
            score = match_score(query_tokens, it["title"])
            flag = "✓" if score >= args.threshold else " "
            print(
                f"  {flag}{i:>2}. {it['lprice']!s:>9}원  매칭{score:.0%}"
                f"  [{it['mallName']}] {it['title']}"
            )

        best, reason = select_best(args.query, items, threshold=args.threshold)
        print(f"\n[단계 2] 선택 결과: {reason}")
        if best:
            print(f"  → {best['lprice']}원  [{best['mallName']}] {best['title']}")
            print(f"     {best['link']}")
        else:
            print("  → 매칭 상품 없음 (결과 엑셀에 빈칸 기록 예정)")

    asyncio.run(_run())


if __name__ == "__main__":
    main()